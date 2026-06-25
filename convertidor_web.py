import os
import math
import re
import io
from datetime import datetime
import pandas as pd
import streamlit as st
import openpyxl

# =====================================================================
# CONFIGURACIÓN DE PÁGINA
# =====================================================================
st.set_page_config(
    page_title="E-Commerce Matrix Connector",
    page_icon="🚀",
    layout="centered",
    initial_sidebar_state="expanded"
)

st.markdown("""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght=400;500;600;700&display=swap');
        html, body, [class*=\"css\"] { font-family: 'Inter', sans-serif; }
        .stFileUploader {
            border: 2px dashed #cbd5e1; border-radius: 12px; padding: 10px;
            background-color: #f8fafc; transition: all 0.3s ease;
        }
        .stFileUploader:hover { border-color: #0284c7; background-color: #f0f9ff; }
        div.stDownloadButton > button {
            background: linear-gradient(135deg, #0284c7, #0369a1) !important;
            color: white !important; border: none !important;
            padding: 12px 24px !important; border-radius: 8px !important;
            font-weight: 600 !important; font-size: 16px !important;
            width: 100% !important;
        }
        div.stDownloadButton > button:hover {
            background: linear-gradient(135deg, #0369a1, #075985) !important;
            transform: translateY(-1px) !important;
        }
    </style>
""", unsafe_allow_html=True)

# =====================================================================
# DICCIONARIOS DE CATEGORÍAS (RUTAS Y PLANTILLAS)
# =====================================================================
CATEGORIAS_FALABELLA = {
    "2316 - Juguetes y juegos / Muñecos de acción no eléctricos":           "juguetes_y_juegos.xlsx",
    "1510 - Juguetes y juegos / Peluches y otras muñecas":                  "peluches.xlsx",
    "956  - Juguetes y juegos / Bloques de construcción (Lego)":            "lego.xlsx",
    "2065 - Juguetes y juegos / Juegos de cartas":                          "cartas.xlsx",
    "449  - Juguetes y juegos / Rompecabezas":                              "rompecabezas.xlsx",
    "1259 - Juguetes y juegos / Juegos de tablero":                         "juego_de_mesa.xlsx",
    "3303 - Ropa y accesorios / Pijamas":                                   "pijamas.xlsx",
    "2898 - Ropa y accesorios / Polos y cassettes":                         "polo.xlsx",
    "463  - Hogar / Ropa de cama":                                          "ropa_de_cama.xlsx",
}

CATEGORIAS_RIPLEY = {
    "Juguetes y Juegos (Figuras Coleccionables)":                          "juguetes_y_juegos.xlsx",
    "Juguetes y Juegos (Bloques y Construcción - LEGO)":                   "ripley_lego.xlsx",
    "Juguetes y Juegos (Juegos de Mesa)":                                  "ripley_juegosdemesa.xlsx",
    "Juguetes y Juegos (Cartas / TCG)":                                    "ripley_tcg.xlsx",
    "Juguetes y Juegos (Peluches)":                                        "ripley_peluches.xlsx",
    "Maletería (Mochilas)":                                                "ripley_mochilas.xlsx",
    "Maletería (Loncheras)":                                               "ripley_lonchera.xlsx",
    "Vestuario Infantil (Tops / Polos)":                                   "ripley_polo.xlsx",
    "Vestuario Infantil (Bottoms / Pantalones)":                           "ripley_pantalones.xlsx",
    "Vestuario Infantil (Conjuntos, Enterizos y Vestidos)":                "ripley_enterizos.xlsx",
    "Relojes de Pulsera":                                                  "ripley_relojes.xlsx",
    "Artículos de Bebé (Bacinicas)":                                       "ripley_bacinicas.xlsx",
    "Tecnología (Smartwatch y Smartband)":                                 "ripley_smartwatch.xlsx",
    "Tecnología (Cargadores)":                                             "ripley_cargadores.xlsx",
    "Calzado (Zapatillas Urbanas)":                                        "ripley_zapatillasurb.xlsx",
    "Deporte y Aventura (Patines)":                                        "ripley_patines.xlsx",
    "Deporte y Aventura (Pelotas y Balones)":                              "ripley_pelotas.xlsx",
    "Hogar (Sábanas y Ropa de Cama)":                                      "ripley_sabanas.xlsx",
}

CATEGORIA_LEGO = "956  - Juguetes y juegos / Bloques de construcción (Lego)"

# =====================================================================
# 1. UTILIDADES COMPARTIDAS
# =====================================================================
_EMOJI_PATTERN = re.compile(
    "["
    "\U0001F000-\U0001FFFF"
    "\U00002000-\U000027FF"
    "\U00002B00-\U00002BFF"
    "\U0000FE00-\U0000FE0F"
    "\U0000FFF0-\U0000FFFF"
    "\u200b-\u200f"
    "\u2028-\u202f"
    "]+",
    flags=re.UNICODE
)

def eliminar_emojis(texto):
    return _EMOJI_PATTERN.sub('', str(texto))

def limpiar_nombre_shopstar(texto):
    if pd.isna(texto):
        return ""
    texto_str = eliminar_emojis(str(texto))
    texto_limpio = re.sub(r'-[^-]+-', '', texto_str)
    texto_limpio = texto_limpio.replace("'", "").replace("-", "")
    return " ".join(texto_limpio.split())

def limpiar_nombre_falabella(texto):
    if pd.isna(texto):
        return ""
    s = eliminar_emojis(str(texto))
    for tag in ['-PREVENTA-', '-OFERTA-', '-LIQUIDACION-']:
        s = s.replace(tag, '')
    return " ".join(s.replace('"', '').replace("'", '').split())

def limpiar_nombre_ripley(texto):
    if pd.isna(texto):
        return ""
    s = eliminar_emojis(str(texto))
    s = s.upper().strip()
    return s[:125]

def limpiar_descripcion(html_texto):
    if pd.isna(html_texto):
        return ""
    texto = str(html_texto)
    texto = texto.replace('\\n', '<br>')
    texto = texto.replace('_x000D_', '\n')
    texto = eliminar_emojis(texto)
    texto = re.sub(r'[ \t]+', ' ', texto).strip()
    return texto

def limpiar_imagenes(val):
    if pd.isna(val):
        return ""
    enlaces = [url.strip() for url in str(val).split(',') if url.strip()]
    return ",".join(enlaces)

def limpiar_sku_falabella(val):
    if pd.isna(val):
        return ""
    return re.sub(r'[^A-Za-z0-9]', '', str(val))

def safe_num(val, default):
    try:
        v = float(val)
        return v if not pd.isna(v) and v > 0 else default
    except (ValueError, TypeError):
        return default

# =====================================================================
# 2. MOTOR SHOPSTAR (Plantilla Local)
# =====================================================================
RUTA_PLANTILLA_SHOPSTAR = os.path.join("plantillas_shopstar", "Plantillashoptar_.xlsx")

def calcular_stock_shopstar(val):
    try:
        stock = float(val)
        if pd.isna(stock) or stock <= 0:
            return 0
        if stock in [1, 2]:
            return 1
        return math.ceil(stock * 0.25)
    except (ValueError, TypeError):
        return 0

def calcular_precio_especial_shopstar(val):
    try:
        precio = float(val)
        if pd.isna(precio) or precio <= 0:
            return 0
        precio_ajustado = precio * 1.26
        return math.ceil((precio_ajustado - 9) / 10) * 10 + 9
    except (ValueError, TypeError):
        return 0

def procesar_logica_shopstar(df_wp, df_marcas_maestro):
    logs = []

    try:
        wb_base = openpyxl.load_workbook(RUTA_PLANTILLA_SHOPSTAR)
    except FileNotFoundError:
        logs.append((f"❌ No se encontró la plantilla '{RUTA_PLANTILLA_SHOPSTAR}'. Asegúrate de tenerla en la carpeta 'plantillas_shopstar/' del repositorio.", "error"))
        return None, None, logs, True

    ws_base = wb_base['PLANTILLA']
    n_cols = ws_base.max_column
    col_nombres = [str(ws_base.cell(row=1, column=c).value or "").strip() for c in range(1, n_cols + 1)]

    columnas_requeridas_wp = [
        'SKU', 'Nombre', 'Descripción', 'Marcas',
        'Inventario', 'Precio normal', 'Imágenes',
        'Peso (kg)', 'Longitud (cm)', 'Anchura (cm)', 'Altura (cm)'
    ]
    columnas_faltantes_wp = [col for col in columnas_requeridas_wp if col not in df_wp.columns]
    if columnas_faltantes_wp:
        logs.append(("❌ ERROR: El CSV de WordPress no tiene la estructura correcta.\nFaltan:\n" + "\n".join([f"  ⚠️ '{c}'" for c in columnas_faltantes_wp]), "error"))
        return None, None, logs, True

    if 'MARCA WP' not in df_marcas_maestro.columns or 'MARCA SS' not in df_marcas_maestro.columns:
        logs.append(("❌ ERROR: La tabla de marcas debe tener los encabezados 'MARCA WP' y 'MARCA SS'.", "error"))
        return None, None, logs, True

    dict_marcas = dict(zip(df_marcas_maestro['MARCA WP'].astype(str).str.strip().str.lower(), df_marcas_maestro['MARCA SS'].astype(str).str.strip()))
    marcas_en_wp = df_wp['Marcas'].dropna().unique()
    marcas_faltantes = [str(m).strip() for m in marcas_en_wp if str(m).strip().lower() not in dict_marcas]
    if marcas_faltantes:
        logs.append(("❌ PROCESO DETENIDO: Marcas nuevas sin registrar en equivalencias:\n" + "\n".join([f"  • {m}" for m in marcas_faltantes]), "error"))
        return None, None, logs, True

    logs.append(("✨ Aplicando transformaciones y limpiando textos...", "info"))
    n = len(df_wp)
    precio_para_especial = pd.to_numeric(df_wp.get('Precio rebajado', pd.Series([0]*n)), errors='coerce').fillna(0)
    precio_normal        = pd.to_numeric(df_wp['Precio normal'], errors='coerce').fillna(0)
    precio_base_especial = precio_para_especial.where(precio_para_especial > 0, precio_normal)

    precios_especiales = precio_base_especial.apply(calcular_precio_especial_shopstar)

    filas_productos = []
    for i in range(n):
        fila = {col: "" for col in col_nombres if col}
        precio_esp = int(precios_especiales.iloc[i])
        precio_base = int(round(precio_esp * 1.5))
        nombre_limpio = limpiar_nombre_shopstar(df_wp['Nombre'].iloc[i])
        sku = str(df_wp['SKU'].iloc[i]).strip() if not pd.isna(df_wp['SKU'].iloc[i]) else ""

        for col in col_nombres:
            if not col: continue
            col_l = col.lower()
            if col == 'Link Imagenes':       fila[col] = limpiar_imagenes(df_wp['Imágenes'].iloc[i])
            elif col == 'Categoria':         fila[col] = '1038-Infantil/Juguetes/Coleccionables'
            elif col == 'Nombre Producto':   fila[col] = nombre_limpio
            elif col == 'Nombre SKU':        fila[col] = nombre_limpio
            elif col == 'SKU':               fila[col] = sku
            elif col == 'Descripcion':       fila[col] = limpiar_descripcion(df_wp['Descripción'].iloc[i])
            elif col == 'Marca':             fila[col] = dict_marcas.get(str(df_wp['Marcas'].iloc[i]).strip().lower(), "")
            elif col == 'Peso':              fila[col] = safe_num(df_wp.get('Peso (kg)', pd.Series([0]*n)).iloc[i], 0)*1000
            elif col == 'Alto':              fila[col] = safe_num(df_wp.get('Altura (cm)', pd.Series([0]*n)).iloc[i], 0)
            elif col == 'Ancho':             fila[col] = safe_num(df_wp.get('Anchura (cm)', pd.Series([0]*n)).iloc[i], 0)
            elif col == 'Largo':             fila[col] = safe_num(df_wp.get('Longitud (cm)', pd.Series([0]*n)).iloc[i], 0)
            elif col == 'Stock':             fila[col] = int(calcular_stock_shopstar(df_wp['Inventario'].iloc[i]))
            elif col == 'Precio Especial':   fila[col] = precio_esp
            elif col == 'Precio Base':       fila[col] = precio_base
            elif col == 'Link':              fila[col] = nombre_limpio.lower().replace(' ', '-') + '-' + sku.lower()
        filas_productos.append(fila)

    df_out = pd.DataFrame(filas_productos, columns=col_nombres)
    logs.append((f"✅ Conversión finalizada con éxito. Se procesaron {len(df_out)} productos.", "success"))
    return df_out, wb_base, logs, False

def generar_excel_shopstar(df_productos, wb_base):
    ws = wb_base['PLANTILLA']
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row: cell.value = None
    for row_idx, row_data in enumerate(df_productos.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value if value != '' else None)
    output = io.BytesIO()
    wb_base.save(output)
    return output.getvalue()


# =====================================================================
# 3. MOTOR FALABELLA (CORREGIDO COMPLETAMENTE)
# =====================================================================
def calcular_sale_price_falabella(val):
    try:
        p = float(val)
        return int(round(p * 1.26 / 10) * 10 - 1) if not pd.isna(p) and p > 0 else 0
    except (ValueError, TypeError):
        return 0

def calcular_stock_falabella(val):
    try:
        s = float(val)
        if pd.isna(s) or s <= 0:
            return 0
        return math.ceil(s * 0.5)
    except (ValueError, TypeError):
        return 0

def procesar_logica_falabella(df_wp, df_marcas_maestro, categoria_sel):
    logs = []
    nombre_archivo = CATEGORIAS_FALABELLA[categoria_sel]
    ruta_plantilla = os.path.join("plantillas_falabella", nombre_archivo)

    try:
        wb_base = openpyxl.load_workbook(ruta_plantilla)
    except FileNotFoundError:
        logs.append((f"❌ No se encontró la plantilla '{ruta_plantilla}'. Asegúrate de tenerla en tu repositorio.", "error"))
        return None, None, logs, True

    ws_base = wb_base['Subir plantilla']
    n_cols = ws_base.max_column
    col_nombres = [str(ws_base.cell(row=4, column=c).value or "").strip() for c in range(1, n_cols + 1)]

    ws_cat = wb_base['Categorías']
    categoria_prim = ""
    for row in ws_cat.iter_rows(max_row=10, max_col=3, values_only=True):
        for cell in row:
            if cell and str(cell).strip() and str(cell).strip() != "PrimaryCategory":
                categoria_prim = str(cell).strip()
                break
        if categoria_prim: break

    columnas_requeridas_wp = ['SKU', 'Nombre', 'Descripción', 'Marcas', 'Inventario', 'Imágenes']
    faltantes = [c for c in columnas_requeridas_wp if c not in df_wp.columns]
    if faltantes:
        logs.append(("❌ ERROR: Faltan columnas en el CSV de WordPress.", "error"))
        return None, None, logs, True

    if 'MARCA WP' not in df_marcas_maestro.columns or 'MARCA FALABELLA' not in df_marcas_maestro.columns:
        logs.append(("❌ ERROR: Equivalencias debe tener 'MARCA WP' y 'MARCA FALABELLA'.", "error"))
        return None, None, logs, True

    dict_marcas = dict(zip(df_marcas_maestro['MARCA WP'].astype(str).str.strip().str.lower(), df_marcas_maestro['MARCA FALABELLA'].astype(str).str.strip()))
    marcas_en_wp = df_wp['Marcas'].dropna().unique()
    marcas_sin_equiv = [str(m).strip() for m in marcas_en_wp if str(m).strip().lower() not in dict_marcas]
    if marcas_sin_equiv:
        logs.append(("❌ PROCESO DETENIDO: Marcas sin equivalencia en 'MARCA FALABELLA':\n" + "\n".join([f"  • {m}" for m in marcas_sin_equiv]), "error"))
        return None, None, logs, True

    logs.append((f"✨ Procesando con plantilla: {nombre_archivo}...", "info"))
    hoy = datetime.now().strftime('%Y-%m-%d')
    n = len(df_wp)

    precio_rebajado = pd.to_numeric(df_wp.get('Precio rebajado', pd.Series([0]*n)), errors='coerce').fillna(0)
    precio_normal   = pd.to_numeric(df_wp.get('Precio normal',   pd.Series([0]*n)), errors='coerce').fillna(0)
    precio_origen   = precio_rebajado.where(precio_rebajado > 0, precio_normal)

    sale_prices = precio_origen.apply(calcular_sale_price_falabella).astype(int)
    list_prices = (sale_prices * 1.5).apply(math.ceil).astype(int)
    stocks      = df_wp['Inventario'].apply(calcular_stock_falabella).astype(int)
    pais_produccion = "Dinamarca" if categoria_sel == CATEGORIA_LEGO else "China"

    col_gtin = 'GTIN, UPC, EAN o ISBN'
    gtin_series = df_wp[col_gtin] if col_gtin in df_wp.columns else pd.Series([''] * n)

    filas_productos = []
    for i in range(n):
        fila = {col: "" for col in col_nombres}
        imgs_raw = df_wp['Imágenes'].iloc[i]
        imgs = [u.strip() for u in str(imgs_raw).split(',') if u.strip()] if not pd.isna(imgs_raw) else []
        imgs_padded = (imgs + [''] * 8)[:8]

        ancho_pkg = safe_num(df_wp.get('Anchura (cm)',  pd.Series([10]*n)).iloc[i], 10)
        largo_pkg = safe_num(df_wp.get('Longitud (cm)', pd.Series([10]*n)).iloc[i], 10)
        alto_pkg  = safe_num(df_wp.get('Altura (cm)',   pd.Series([10]*n)).iloc[i], 10)
        peso_pkg  = safe_num(df_wp.get('Peso (kg)',     pd.Series([0.5]*n)).iloc[i], 0.5)

        img_idx = 0
        for col in col_nombres:
            if not col: continue
            
            # === TODO EL BLOQUE EVALUADO CORRECTAMENTE CON .lower() ===
            if "nombre #" in col.lower(): 
                fila[col] = limpiar_nombre_falabella(df_wp['Nombre'].iloc[i])
            elif "marca #" in col.lower(): 
                fila[col] = dict_marcas.get(str(df_wp['Marcas'].iloc[i]).strip().lower(), "")
            elif "descripción #" in col.lower() or "descripcion #" in col.lower(): 
                fila[col] = limpiar_descripcion(df_wp['Descripción'].iloc[i])
            elif "categoría primaria" in col.lower() or "primarycategory" in col.lower(): 
                fila[col] = categoria_prim
            elif "país de producción" in col.lower() or "productioncountry" in col.lower(): 
                fila[col] = pais_produccion
            elif "sku del vendedor" in col.lower() or "seller_sku" in col.lower(): 
                fila[col] = limpiar_sku_falabella(df_wp['SKU'].iloc[i])
            elif "código de barras" in col.lower() or "barcode" in col.lower() or "gtin" in col.lower():
                try:
                    fila[col] = int(float(gtin_series.iloc[i])) if not pd.isna(gtin_series.iloc[i]) and str(gtin_series.iloc[i]).strip() not in ('', 'nan') else ""
                except (ValueError, TypeError):
                    fila[col] = ""
            elif "variación #" in col.lower() or "variacion #" in col.lower(): 
                fila[col] = "..."
            elif "quantity" in col.lower() or "stock" in col.lower(): 
                fila[col] = int(stocks.iloc[i])
            elif "price" in col.lower() and "sale" not in col.lower() and "discount" not in col.lower(): 
                fila[col] = int(list_prices.iloc[i])
            elif "sale-price" in col.lower() or "saleprice" in col.lower(): 
                fila[col] = int(sale_prices.iloc[i])
            elif "salestartdate" in col.lower() or "fecha inicio" in col.lower(): 
                fila[col] = hoy
            elif "saleenddate" in col.lower() or "fecha fin" in col.lower(): 
                fila[col] = "2050-01-01"
            elif "condición del producto" in col.lower() or "productcondition" in col.lower(): 
                fila[col] = "Nuevo"
            elif "grupodeedad" in col.lower() or "grupo de edad" in col.lower(): 
                fila[col] = "Todas las edades"
            elif "piezaspequenas" in col.lower() or "piezas pequeñas" in col.lower(): 
                fila[col] = "Sí"
            elif "caracteristicasdesalud" in col.lower() or "características de salud" in col.lower(): 
                fila[col] = "Sin BPA"
            elif "weightoftheproduct" in col.lower():
                fila[col] = f"{peso_pkg} kg"
            elif "weight" in col.lower() or "peso del paquete" in col.lower(): 
                fila[col] = peso_pkg
            elif "ancho del paquete" in col.lower() or "width" in col.lower(): 
                fila[col] = ancho_pkg
            elif "largo del paquete" in col.lower() or "length" in col.lower(): 
                fila[col] = largo_pkg
            elif "alto del paquete" in col.lower() or "height" in col.lower(): 
                fila[col] = alto_pkg
            elif "color #" in col.lower():
                fila[col] = "MULTICOLOR"
            elif "alto #" in col.lower():
                fila[col] = f"{alto_pkg} cm"
            elif "ancho #" in col.lower():
                fila[col] = f"{ancho_pkg} cm"
            elif "largo #" in col.lower():
                fila[col] = f"{largo_pkg} cm"
            elif "dimensiones" in col.lower():
                fila[col] = f"{alto_pkg} cm x {largo_pkg} cm x {ancho_pkg} cm"
            elif "imagen principal" in col.lower() or "imagen" in col.lower() or "image" in col.lower():
                if img_idx < 8:
                    fila[col] = imgs_padded[img_idx]
                    img_idx += 1
        filas_productos.append(fila)

    df_productos = pd.DataFrame(filas_productos, columns=col_nombres)
    logs.append((f"✅ Falabella finalizado con éxito.", "success"))
    return df_productos, wb_base, logs, False

def _convert_to_shared_strings(xlsx_bytes):
    """
    Post-procesa el xlsx generado por openpyxl para compatibilidad estricta con Falabella:
    1. Convierte inlineStr → sharedStrings
    2. Elimina celdas vacías con t="n"
    3. Quita t="n" de celdas numéricas con valor
    4. Corrige rutas absolutas → relativas en workbook.xml.rels  ← FIX CRÍTICO
    5. Agrega declaración XML a los archivos que la necesitan
    """
    import zipfile as _zf, re as _re
    src = _zf.ZipFile(io.BytesIO(xlsx_bytes))
    dst_buf = io.BytesIO()
    dst = _zf.ZipFile(dst_buf, 'w', _zf.ZIP_DEFLATED)
    shared_strings = []
    ss_index = {}

    def get_idx(val):
        if val not in ss_index:
            ss_index[val] = len(shared_strings)
            shared_strings.append(val)
        return ss_index[val]

    pattern = r'<c([^>]*)>\s*<is>\s*<t(?:[^>]*)>(.*?)</t>\s*</is>\s*</c>'

    # ── Paso 1: procesar worksheets ──────────────────────────────────
    sheet_xmls = {}
    for name in src.namelist():
        if _re.match(r'xl/worksheets/sheet\d+\.xml$', name):
            xml = src.read(name).decode('utf-8').lstrip('\ufeff')
            xml = _re.sub(r'^\s*<\?xml[^?]*\?>\s*', '', xml)

            def replace_inline(m):
                attrs = _re.sub(r'\s*t="inlineStr"', '', m.group(1))
                return f'<c{attrs} t="s"><v>{get_idx(m.group(2))}</v></c>'
            xml = _re.sub(pattern, replace_inline, xml, flags=_re.DOTALL)
            xml = _re.sub(r'<c[^>]* t="n"[^>]*>\s*</c>', '', xml)
            xml = _re.sub(r'(<c\b[^>]*?) t="n"([^>]*>)', r'\1\2', xml)
            xml = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n' + xml
            sheet_xmls[name] = xml.encode('utf-8')

    # ── Paso 2: escribir todos los archivos ──────────────────────────
    for name in src.namelist():
        if name in sheet_xmls:
            dst.writestr(name, sheet_xmls[name])

        elif name == 'xl/sharedStrings.xml':
            pass  # regenerado al final

        elif name == 'xl/_rels/workbook.xml.rels':
            rels = src.read(name).decode('utf-8').lstrip('\ufeff')
            rels = _re.sub(r'^\s*<\?xml[^?]*\?>\s*', '', rels)
            # FIX: rutas absolutas → relativas (openpyxl a veces las escribe absolutas)
            rels = _re.sub(r'Target="/xl/worksheets/', 'Target="worksheets/', rels)
            rels = _re.sub(r'Target="/xl/styles\.xml"', 'Target="styles.xml"', rels)
            rels = _re.sub(r'Target="/xl/theme/', 'Target="theme/', rels)
            rels = _re.sub(r'Target="/xl/sharedStrings\.xml"', 'Target="sharedStrings.xml"', rels)
            if 'sharedStrings' not in rels:
                rels = rels.replace(
                    '</Relationships>',
                    '<Relationship Id="rIdSS" '
                    'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" '
                    'Target="sharedStrings.xml"/></Relationships>'
                )
            rels = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n' + rels
            dst.writestr(name, rels.encode('utf-8'))

        elif name == '[Content_Types].xml':
            ct = src.read(name).decode('utf-8').lstrip('\ufeff')
            ct = _re.sub(r'^\s*<\?xml[^?]*\?>\s*', '', ct)
            if 'sharedStrings' not in ct:
                ct = ct.replace(
                    '</Types>',
                    '<Override PartName="/xl/sharedStrings.xml" '
                    'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>'
                    '</Types>'
                )
            ct = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n' + ct
            dst.writestr(name, ct.encode('utf-8'))

        elif name == 'xl/workbook.xml':
            wb = src.read(name).decode('utf-8').lstrip('\ufeff')
            wb = _re.sub(r'^\s*<\?xml[^?]*\?>\s*', '', wb)
            wb = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n' + wb
            dst.writestr(name, wb.encode('utf-8'))

        else:
            dst.writestr(name, src.read(name))

    # ── Paso 3: sharedStrings.xml ────────────────────────────────────
    if shared_strings:
        n = len(shared_strings)
        ss_xml = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
        ss_xml += f'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="{n}" uniqueCount="{n}">'
        for s in shared_strings:
            esc = _re.sub(r'&(?!(?:amp|lt|gt|quot|apos|#\d+|#x[\da-fA-F]+);)', '&amp;', s)
            ss_xml += f'<si><t xml:space="preserve">{esc}</t></si>'
        ss_xml += '</sst>'
        dst.writestr('xl/sharedStrings.xml', ss_xml.encode('utf-8'))

    src.close(); dst.close(); dst_buf.seek(0)
    return dst_buf.read()

def generar_excel_falabella(df_productos, wb_base):
    ws = wb_base['Subir plantilla']
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
        for cell in row: cell.value = None
    col_barras = next((c for c in range(1, ws.max_column + 1)
                       if 'barras' in str(ws.cell(row=4, column=c).value or '').lower()), None)
    for row_idx, row_data in enumerate(df_productos.itertuples(index=False), start=5):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value if value != '' else None)
            if col_idx == col_barras and isinstance(value, int):
                cell.number_format = '0'
    output = io.BytesIO()
    wb_base.save(output)
    return _convert_to_shared_strings(output.getvalue())


# =====================================================================
# 4. MOTOR RIPLEY (Lectura desde Plantilla Local)
# =====================================================================
def calcular_stock_ripley(val):
    try:
        stock = float(val)
        if pd.isna(stock) or stock <= 0:
            return 1
        return max(1, math.ceil(stock * 0.25))
    except (ValueError, TypeError):
        return 1

def calcular_discount_price_ripley(val):
    try:
        precio = float(val)
        if pd.isna(precio) or precio <= 0:
            return 0
        precio_ajustado = precio * 1.26
        return math.ceil((precio_ajustado - 9) / 10) * 10 + 9
    except (ValueError, TypeError):
        return 0

def procesar_logica_ripley(df_wp, df_marcas_maestro, categoria_sel):
    logs = []
    nombre_archivo = CATEGORIAS_RIPLEY[categoria_sel]
    ruta_plantilla = os.path.join("plantillas_ripley", nombre_archivo)

    try:
        wb_base = openpyxl.load_workbook(ruta_plantilla)
    except FileNotFoundError:
        logs.append((f"❌ No se encontró la plantilla de Ripley '{ruta_plantilla}'. Asegúrate de que exista la carpeta y el archivo.", "error"))
        return None, None, logs, True

    ws_data = wb_base['Data']
    n_cols = ws_data.max_column
    col_nombres = [str(ws_data.cell(row=2, column=c).value or "").strip() for c in range(1, n_cols + 1)]

    # Leer la categoría exacta desde ReferenceData (fila 2, columna 1)
    ws_ref = wb_base['ReferenceData']
    categoria_ripley = str(ws_ref.cell(row=2, column=1).value or "").strip()

    columnas_requeridas_wp = ['SKU', 'Nombre', 'Descripción', 'Marcas', 'Inventario', 'Imágenes']
    faltantes = [c for c in columnas_requeridas_wp if c not in df_wp.columns]
    if faltantes:
        logs.append(("❌ ERROR: Al CSV de WordPress le faltan columnas requeridas para Ripley:\n" + "\n".join([f"  ⚠️ '{c}'" for c in faltantes]), "error"))
        return None, None, logs, True

    if 'MARCA WP' not in df_marcas_maestro.columns or 'MARCA RIPLEY' not in df_marcas_maestro.columns:
        logs.append(("❌ ERROR: La tabla de equivalencias debe incluir 'MARCA WP' y 'MARCA RIPLEY'.", "error"))
        return None, None, logs, True

    dict_marcas = dict(zip(df_marcas_maestro['MARCA WP'].astype(str).str.strip().str.lower(), df_marcas_maestro['MARCA RIPLEY'].astype(str).str.strip()))
    marcas_en_wp = df_wp['Marcas'].dropna().unique()
    marcas_faltantes = [str(m).strip() for m in marcas_en_wp if str(m).strip().lower() not in dict_marcas]
    if marcas_faltantes:
        logs.append(("❌ PROCESO DETENIDO: Marcas sin equivalencia en 'MARCA RIPLEY':\n" + "\n".join([f"  • {m}" for m in marcas_faltantes]), "error"))
        return None, None, logs, True

    logs.append((f"✨ Leyendo estructura e inyectando datos desde plantilla: {ruta_plantilla}...", "info"))
    hoy_iso = datetime.now().strftime('%Y-%m-%dT%H:%M:%S.000-05:00')
    fecha_fin_iso = '2050-06-09T23:59:59.000-05:00'
    n = len(df_wp)

    precio_rebajado = pd.to_numeric(df_wp.get('Precio rebajado', pd.Series([0]*n)), errors='coerce').fillna(0)
    precio_normal   = pd.to_numeric(df_wp.get('Precio normal',   pd.Series([0]*n)), errors='coerce').fillna(0)
    precio_origen   = precio_rebajado.where(precio_rebajado > 0, precio_normal)

    filas_productos = []
    for i in range(n):
        fila = {col: "" for col in col_nombres if col}
        imgs_raw = df_wp['Imágenes'].iloc[i]
        imgs = [u.strip() for u in str(imgs_raw).split(',') if u.strip()] if not pd.isna(imgs_raw) else []

        disc_p = calcular_discount_price_ripley(precio_origen.iloc[i])
        list_p = int(math.ceil(disc_p * 1.5))
        stock_calc = int(calcular_stock_ripley(df_wp['Inventario'].iloc[i]))

        largo = safe_num(df_wp.get('Longitud (cm)', pd.Series([10]*n)).iloc[i], 10)
        ancho = safe_num(df_wp.get('Anchura (cm)',  pd.Series([10]*n)).iloc[i], 10)
        alto  = safe_num(df_wp.get('Altura (cm)',   pd.Series([10]*n)).iloc[i], 10)
        peso  = safe_num(df_wp.get('Peso (kg)',     pd.Series([0.5]*n)).iloc[i], 0.5)

        for col in col_nombres:
            if not col: continue
            if col == 'categoria': fila[col] = categoria_ripley
            elif col == 'sku_seller': fila[col] = str(df_wp['SKU'].iloc[i]).strip()
            elif col == 'nombre': fila[col] = limpiar_nombre_ripley(df_wp['Nombre'].iloc[i])
            elif col == 'ShortDescription': fila[col] = 'Encuentra los mejores productos de Raymi Store en Ripley.com'
            elif col == 'descripcion': fila[col] = limpiar_descripcion(df_wp['Descripción'].iloc[i])
            elif col == 'marca': fila[col] = dict_marcas.get(str(df_wp['Marcas'].iloc[i]).strip().lower(), "").upper()
            elif col == 'imagen': fila[col] = imgs[0] if len(imgs) > 0 else ""
            elif col == 'thumbnail': fila[col] = imgs[0] if len(imgs) > 0 else ""
            elif col.startswith('imagen') and col[6:].isdigit():
                idx = int(col[6:])
                if len(imgs) >= idx: fila[col] = imgs[idx-1]
            elif col == 'largo_empaque': fila[col] = largo
            elif col == 'ancho_empaque': fila[col] = ancho
            elif col == 'alto_empaque': fila[col] = alto
            elif col == 'peso_empaque': fila[col] = peso
            elif col == 'tipo_product': fila[col] = 'Coleccionables'
            elif col == 'color_80': fila[col] = 'Multicolor'
            elif col == 'edad_recomendada': fila[col] = 'Todas las edades'
            elif col == 'alto_cm': fila[col] = alto
            elif col == 'ancho_cm': fila[col] = ancho
            elif col == 'largo_cm': fila[col] = largo
            elif col == 'sku': fila[col] = str(df_wp['SKU'].iloc[i]).strip()
            elif col == 'product-id': fila[col] = str(df_wp['SKU'].iloc[i]).strip()
            elif col == 'product-id-type': fila[col] = 'SHOP_SKU'
            elif col == 'discount-price': fila[col] = int(disc_p)
            elif col == 'price': fila[col] = int(list_p)
            elif col == 'quantity': fila[col] = stock_calc
            elif col == 'state': fila[col] = 'Nuevo'
            elif col == 'available-start-date': fila[col] = hoy_iso
            elif col == 'available-end-date': fila[col] = fecha_fin_iso
            
        filas_productos.append(fila)

    df_productos = pd.DataFrame(filas_productos, columns=col_nombres)
    logs.append((f"✅ Ripley procesado correctamente. {len(df_productos)} filas listas.", "success"))
    return df_productos, wb_base, logs, False

def generar_excel_ripley(df_productos, wb_base):
    ws = wb_base['Data']
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row: cell.value = None

    # Identificar índices de columnas de fecha para forzar formato texto
    n_cols = ws.max_column
    nombres_tecnicos = [ws.cell(row=2, column=c).value or "" for c in range(1, n_cols + 1)]
    col_fechas = {i + 1 for i, t in enumerate(nombres_tecnicos)
                  if t in ('available-start-date', 'available-end-date',
                           'discount-start-date', 'discount-end-date')}

    for row_idx, row_data in enumerate(df_productos.itertuples(index=False), start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value if value != '' else None)
            if col_idx in col_fechas and value:
                cell.number_format = '@'  # formato texto, evita que Excel reinterprete el ISO string

    output = io.BytesIO()
    wb_base.save(output)
    return output.getvalue()


# =====================================================================
# 5. CARGADORES DE ARCHIVOS
# =====================================================================
def cargar_maestro(file_obj):
    if file_obj is None: return None
    name = file_obj.name.lower()
    if name.endswith('.xlsx'):
        try: return pd.read_excel(file_obj, sheet_name='MARCAS')
        except Exception:
            file_obj.seek(0)
            return pd.read_excel(file_obj, sheet_name=0)
    return pd.read_csv(file_obj, sep=',', dtype=str)

def cargar_wp(file_obj):
    if file_obj is None: return None
    name = file_obj.name.lower()
    df = pd.read_excel(file_obj, dtype=str) if name.endswith('.xlsx') else pd.read_csv(file_obj, sep=',', dtype=str)
    for col in ['Inventario', 'Precio normal', 'Precio rebajado', 'Peso (kg)', 'Longitud (cm)', 'Anchura (cm)', 'Altura (cm)']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    return df


# =====================================================================
# 6. INTERFAZ DE USUARIO (STREAMLIT)
# =====================================================================
st.markdown("""
    <div style="background: linear-gradient(135deg, #1e293b, #0f172a); padding: 25px;
                border-radius: 12px; margin-bottom: 25px; text-align: center; color: white;">
        <h1 style="margin: 0; font-size: 28px; font-weight: 700; font-family: 'Inter', sans-serif;">
            E-Commerce Matrix Connector v5.4
        </h1>
        <p style="margin: 8px 0 0 0; color: #94a3b8; font-size: 15px;">
            WordPress → Shopstar &nbsp;|&nbsp; Falabella (Plantilla Local) &nbsp;|&nbsp; Ripley (Plantilla Local)
        </p>
    </div>
""", unsafe_allow_html=True)

with st.sidebar:
    st.image("https://img.icons8.com/clouds/100/null/data-configuration.png", width=100)
    st.markdown("### Carpetas de Repositorio")
    st.markdown("""
    * Shopstar lee de: `plantillas_shopstar/`
    * Falabella lee de: `plantillas_falabella/`
    * Ripley lee de: `plantillas_ripley/`
    """)
    st.divider()

    st.markdown("##### 🗂️ Categoría Falabella")
    categoria_falabella = st.selectbox(
        "Lote Falabella:", options=list(CATEGORIAS_FALABELLA.keys()), key="cat_fal"
    )
    st.divider()

    st.markdown("##### 🎫 Categoría Ripley")
    categoria_ripley = st.selectbox(
        "Lote Ripley:", options=list(CATEGORIAS_RIPLEY.keys()), key="cat_rip"
    )
    st.divider()
    st.info("💡 **Calford Import — v5.4**\nShopstar migrado a plantilla local.")

col1, col2 = st.columns(2)
with col1:
    st.markdown("##### 📁 1. Tabla de Equivalencias de Marcas")
    file_maestro = st.file_uploader("Excel de marcas (.xlsx / .csv)", type=["xlsx", "csv"], key="maestro")
with col2:
    st.markdown("##### 📝 2. Catálogo WordPress")
    file_wp = st.file_uploader("Archivo WordPress (.csv / .xlsx)", type=["csv", "xlsx"], key="wordpress")

st.divider()

col_btn1, col_btn2, col_btn3 = st.columns(3)
btn_shopstar  = col_btn1.button("📦 Generar Formato Shopstar",  use_container_width=True)
btn_falabella = col_btn2.button("🔥 Generar Formato Falabella", use_container_width=True)
btn_ripley    = col_btn3.button("🎫 Generar Formato Ripley",    use_container_width=True)

if btn_shopstar:
    if not file_maestro or not file_wp:
        st.warning("⚠️ Sube ambos archivos antes de generar.")
    else:
        st.markdown("##### ⚙️ Registro de Actividad — Shopstar:")
        with st.spinner("Procesando..."):
            try:
                df_marcas  = cargar_maestro(file_maestro)
                df_wp_data = cargar_wp(file_wp)
                df_resultado, wb_base, logs, has_error = procesar_logica_shopstar(df_wp_data, df_marcas)

                for msg, tipo in logs:
                    if tipo == "error": st.error(msg)
                    elif tipo == "success": st.success(msg)
                    else: st.info(msg)

                if not has_error and df_resultado is not None:
                    st.dataframe(df_resultado.head(5), use_container_width=True)
                    excel_data = generar_excel_shopstar(df_resultado, wb_base)
                    st.download_button(
                        label="🚀 Descargar Plantilla Shopstar (.xlsx)",
                        data=excel_data,
                        file_name=f"Plantilla_Shopstar_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            except Exception as e: st.error(f"❌ ERROR CRÍTICO SHOPSTAR: {str(e)}")

if btn_falabella:
    if not file_maestro or not file_wp:
        st.warning("⚠️ Sube ambos archivos antes de generar.")
    else:
        st.markdown(f"##### ⚙️ Registro de Actividad — Falabella:")
        with st.spinner("Procesando..."):
            try:
                file_maestro.seek(0); file_wp.seek(0)
                df_marcas  = cargar_maestro(file_maestro)
                df_wp_data = cargar_wp(file_wp)
                df_productos, wb_base, logs, has_error = procesar_logica_falabella(df_wp_data, df_marcas, categoria_falabella)

                for msg, tipo in logs:
                    if tipo == "error": st.error(msg)
                    elif tipo == "success": st.success(msg)
                    else: st.info(msg)

                if not has_error and df_productos is not None:
                    st.dataframe(df_productos.head(5), use_container_width=True)
                    excel_data = generar_excel_falabella(df_productos, wb_base)
                    st.download_button(
                        label="🔥 Descargar Plantilla Falabella (.xlsx)",
                        data=excel_data,
                        file_name=f"Plantilla_Falabella_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            except Exception as e: st.error(f"❌ ERROR CRÍTICO FALABELLA: {str(e)}")

if btn_ripley:
    if not file_maestro or not file_wp:
        st.warning("⚠️ Sube ambos archivos antes de generar.")
    else:
        st.markdown("##### ⚙️ Registro de Actividad — Ripley Coleccionables:")
        with st.spinner("Procesando catálogo Ripley local..."):
            try:
                file_maestro.seek(0); file_wp.seek(0)
                df_marcas  = cargar_maestro(file_maestro)
                df_wp_data = cargar_wp(file_wp)
                
                df_productos_ripley, wb_base_ripley, logs, has_error = procesar_logica_ripley(df_wp_data, df_marcas, categoria_ripley)

                for msg, tipo in logs:
                    if tipo == "error": st.error(msg)
                    elif tipo == "success": st.success(msg)
                    else: st.info(msg)

                if not has_error and df_productos_ripley is not None:
                    st.markdown("##### 🔍 Vista Previa Ripley (Estructura Plantilla):")
                    st.dataframe(df_productos_ripley.head(5), use_container_width=True)

                    ripley_excel = generar_excel_ripley(df_productos_ripley, wb_base_ripley)
                    st.markdown("##### 📥 Listo para descargar:")
                    st.download_button(
                        label="🎫 Descargar Plantilla Ripley (.xlsx)",
                        data=ripley_excel,
                        file_name=f"Plantilla_Ripley_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            except Exception as e:
                st.error(f"❌ ERROR CRÍTICO AL PROCESAR RIPLEY: {str(e)}")

if not file_maestro and not file_wp:
    st.info("⬆️ Sube los dos archivos requeridos para activar las opciones de conversión.")